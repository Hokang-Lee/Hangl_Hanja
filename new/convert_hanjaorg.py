# -*- coding: utf-8 -*-
"""
Korean to Hanja converter for Word (.docx) files
- Reads main dictionary from Excel Sheet1 (col A=Korean, col C=Hanja)
- Reads Kyuji->Shinji table from Excel Sheet2 (col A=old, col B=new)
- Handles cross-paragraph and soft-return splits
- Converts all fonts to Meiryo UI after conversion
- Subprocess-safe: opens/closes Word file via python-docx only

Usage:
    python convert_hanja.py "input.docx"
    python convert_hanja.py "input.docx" "output.docx"
    python convert_hanja.py "input.docx" "output.docx" "dict.xlsm"

Requirements:
    pip install python-docx lxml openpyxl
"""
import sys, os
from docx import Document
from docx.shared import Pt
from lxml import etree

NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
FONT_NAME = "Meiryo UI"

# ---------------------------------------------------------------
# Dictionary loader
# ---------------------------------------------------------------
def load_dicts(excel_path):
    """Load main dict (Sheet1) and kyuji table (Sheet2) from Excel."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Sheet1: A=Korean, C=Hanja
    ws1 = wb.worksheets[0]
    main_dict = []
    for r in ws1.iter_rows(min_row=2, values_only=True):
        k = str(r[0]).strip() if r[0] else ""
        h = str(r[2]).strip() if r[2] else ""
        if k and h and k != h and k != "None" and h != "None":
            main_dict.append((k, h))
    main_dict.sort(key=lambda x: -len(x[0]))

    # Sheet2: A=旧字, B=当用漢字
    kyuji = []
    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
        for r in ws2.iter_rows(min_row=2, values_only=True):
            old = str(r[0]).strip() if r[0] else ""
            new = str(r[1]).strip() if r[1] else ""
            if old and new and old != new:
                kyuji.append((old, new))

    wb.close()
    print(f"Main dict: {len(main_dict)} entries")
    print(f"Kyuji table: {len(kyuji)} entries")
    return main_dict, kyuji

# ---------------------------------------------------------------
# XML helpers
# ---------------------------------------------------------------
def get_segments(para_elem):
    segs = []
    for child in para_elem.iter():
        tag = child.tag.split("}")[1] if "}" in child.tag else child.tag
        if tag == "t":
            segs.append([child.text or "", child, False])
        elif tag == "br":
            br_type = child.get(f"{{{NS}}}type", "textWrapping")
            if br_type in ("", "textWrapping"):
                segs.append(["\x0b", child, True])
    return segs

def apply_to_segments(segs, new_full):
    pos = 0
    for seg in segs:
        if seg[2]:
            pos += 1
            continue
        old_len = len(seg[0])
        seg[1].text = new_full[pos:pos + old_len]
        pos += old_len

# ---------------------------------------------------------------
# Replacement logic
# ---------------------------------------------------------------
def replace_text(text, dictionary):
    """Placeholder-based replacement (prevents chain conversion)."""
    placeholders = {}
    result = text
    idx = 0
    for korean, hanja in dictionary:
        if korean in result:
            ph = f"\uE002{idx:04d}\uE003"
            placeholders[ph] = hanja
            result = result.replace(korean, ph)
            idx += 1
    for ph, hanja in placeholders.items():
        result = result.replace(ph, hanja)
    return result

def replace_kyuji(text, kyuji_table):
    """Replace old kanji with standard kanji (1-char, simple replace)."""
    for old, new in kyuji_table:
        text = text.replace(old, new)
    return text

def process_paragraphs(paras_elems, dictionary, kyuji_table):
    """Process paragraph elements: main dict + cross-para + kyuji."""
    changed = 0

    # Collect segments and full texts
    para_segs  = []
    para_texts = []
    for para in paras_elems:
        segs = get_segments(para)
        full = "".join(s[0] for s in segs)
        para_segs.append(segs)
        para_texts.append(full)

    # Step 1: Single-paragraph main dict replacement
    new_texts = [replace_text(t, dictionary) for t in para_texts]

    # Step 2: Cross-paragraph replacement
    # IMPORTANT: Search using ORIGINAL para_texts (before Step1 conversion)
    # to avoid missing words split across paragraphs where part2 was already converted
    for i in range(len(para_texts) - 1):
        orig1 = para_texts[i]
        orig2 = para_texts[i + 1]
        combined_orig = orig1 + "\x00" + orig2
        for korean, hanja in dictionary:
            if len(korean) < 2:
                continue
            for j in range(1, len(korean)):
                part1 = korean[:j]
                part2 = korean[j:]
                split_k = part1 + "\x00" + part2
                if split_k not in combined_orig:
                    continue
                pos = combined_orig.find(split_k)
                boundary = len(orig1)
                if pos < boundary <= pos + len(split_k):
                    # charBefore check: skip if preceded by Hangul or CJK
                    if pos > 0:
                        cb = combined_orig[pos - 1]
                        if ('가' <= cb <= '힣') or ('一' <= cb <= '鿿'):
                            continue
                    if orig1.endswith(part1) and orig2.startswith(part2):
                        hpart1 = hanja[:j] if j < len(hanja) else hanja
                        hpart2 = hanja[j:] if j < len(hanja) else ""
                        # Apply to new_texts (already Step1-converted)
                        # Replace the tail of new_texts[i] that corresponds to part1
                        if new_texts[i].endswith(part1):
                            new_texts[i] = new_texts[i][:-len(part1)] + hpart1
                        # Replace the head of new_texts[i+1] that corresponds to part2
                        # (part2 may already be converted in new_texts[i+1])
                        converted_part2 = replace_text(part2, dictionary)
                        if new_texts[i + 1].startswith(converted_part2):
                            new_texts[i + 1] = hpart2 + new_texts[i + 1][len(converted_part2):]
                        elif new_texts[i + 1].startswith(part2):
                            new_texts[i + 1] = hpart2 + new_texts[i + 1][len(part2):]

    # Step 3: Kyuji -> Shinji replacement
    new_texts = [replace_kyuji(t, kyuji_table) for t in new_texts]

    # Apply changes
    for segs, old, new in zip(para_segs, para_texts, new_texts):
        if old != new:
            apply_to_segments(segs, new)
            changed += 1

    return changed


# ---------------------------------------------------------------
# Font size conversion: 16pt numbers -> 14pt
# ---------------------------------------------------------------
def convert_font_size(elem):
    """Change font size from 16pt to 14pt for numeric runs only."""
    # In OOXML, font size is stored in half-points (1pt = 2 units)
    # 16pt = 32 half-points, 14pt = 28 half-points
    TARGET_SZ  = "32"   # 16pt in half-points
    NEW_SZ     = "28"   # 14pt in half-points

    changed = 0
    for r in elem.iter(f"{{{NS}}}r"):
        # Check if this run has font size 16pt
        rpr = r.find(f"{{{NS}}}rPr")
        if rpr is None:
            continue
        sz = rpr.find(f"{{{NS}}}sz")
        if sz is None:
            continue
        if sz.get(f"{{{NS}}}val") != TARGET_SZ:
            continue

        # Check if run text consists only of digits/spaces/punctuation
        t = r.find(f"{{{NS}}}t")
        if t is None or not t.text:
            continue
        # Only convert if text contains at least one digit
        if not any(c.isdigit() for c in t.text):
            continue

        # Change sz val
        sz.set(f"{{{NS}}}val", NEW_SZ)
        # Also change szCs (complex script size) if present
        szCs = rpr.find(f"{{{NS}}}szCs")
        if szCs is not None and szCs.get(f"{{{NS}}}val") == TARGET_SZ:
            szCs.set(f"{{{NS}}}val", NEW_SZ)
        changed += 1
    return changed

# ---------------------------------------------------------------
# Font conversion
# ---------------------------------------------------------------
def set_font_all(elem, font_name):
    """Set all rFonts in element to font_name (ascii, eastAsia, hAnsi, cs)."""
    for rpr in elem.iter(f"{{{NS}}}rPr"):
        # Remove existing rFonts
        for rf in rpr.findall(f"{{{NS}}}rFonts"):
            rpr.remove(rf)
        # Add new rFonts
        rf = etree.SubElement(rpr, f"{{{NS}}}rFonts")
        rf.set(f"{{{NS}}}ascii",    font_name)
        rf.set(f"{{{NS}}}eastAsia", font_name)
        rf.set(f"{{{NS}}}hAnsi",    font_name)
        rf.set(f"{{{NS}}}cs",       font_name)
        rpr.insert(0, rf)

    # Also set default run properties at paragraph level
    for ppr in elem.iter(f"{{{NS}}}pPr"):
        rpr = ppr.find(f"{{{NS}}}rPr")
        if rpr is None:
            rpr = etree.SubElement(ppr, f"{{{NS}}}rPr")
        for rf in rpr.findall(f"{{{NS}}}rFonts"):
            rpr.remove(rf)
        rf = etree.SubElement(rpr, f"{{{NS}}}rFonts")
        rf.set(f"{{{NS}}}ascii",    font_name)
        rf.set(f"{{{NS}}}eastAsia", font_name)
        rf.set(f"{{{NS}}}hAnsi",    font_name)
        rf.set(f"{{{NS}}}cs",       font_name)
        rpr.insert(0, rf)

# ---------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------

# ---------------------------------------------------------------
# PDF conversion
# ---------------------------------------------------------------
def convert_to_pdf(docx_path, pdf_path):
    """Convert docx to PDF using docx2pdf or LibreOffice."""

    # Method 1: docx2pdf (uses installed Word on Windows/Mac)
    try:
        from docx2pdf import convert as _pdf_convert
        print(f"Converting to PDF (docx2pdf)...")
        _pdf_convert(docx_path, pdf_path)
        print(f"PDF:    {pdf_path}")
        return
    except ImportError:
        pass
    except Exception as e:
        print(f"docx2pdf failed: {e}")

    # Method 2: LibreOffice
    import subprocess, shutil
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice:
        try:
            print(f"Converting to PDF (LibreOffice)...")
            out_dir = os.path.dirname(os.path.abspath(pdf_path))
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf",
                 "--outdir", out_dir, docx_path],
                check=True, timeout=60
            )
            # LibreOffice outputs <basename>.pdf in out_dir
            base = os.path.splitext(os.path.basename(docx_path))[0] + ".pdf"
            lo_out = os.path.join(out_dir, base)
            if lo_out != pdf_path and os.path.exists(lo_out):
                os.replace(lo_out, pdf_path)
            print(f"PDF:    {pdf_path}")
            return
        except Exception as e:
            print(f"LibreOffice failed: {e}")

    print("WARNING: PDF conversion skipped.")
    print("  To enable PDF output, run one of:")
    print("    pip install docx2pdf       (requires Microsoft Word)")
    print("    Install LibreOffice        (free)")

def find_dict_file(script_dir):
    candidates = []
    for name in os.listdir(script_dir):
        full = os.path.join(script_dir, name)
        if not os.path.isfile(full):
            continue
        low = name.lower()
        if low.endswith((".xlsx", ".xlsm")):
            candidates.append((0 if "convertalltext" in low else 1, full))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]

def convert(input_path, output_path=None, dict_path=None):
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = base + "_hanja" + ext

    if dict_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        dict_path  = find_dict_file(script_dir)
        if dict_path is None:
            print("ERROR: No Excel dictionary found in script folder.")
            sys.exit(1)

    print(f"Dictionary: {dict_path}")
    main_dict, kyuji_table = load_dicts(dict_path)

    print(f"Input:  {input_path}")
    doc = Document(input_path)

    # --- Body ---
    body_paras = doc.element.body.findall(f".//{{{NS}}}p")
    body_changed = process_paragraphs(body_paras, main_dict, kyuji_table)

    # --- Headers / Footers ---
    hf_changed = 0
    for section in doc.sections:
        for hf in [section.header, section.footer,
                   section.even_page_header, section.even_page_footer,
                   section.first_page_header, section.first_page_footer]:
            try:
                hf_paras = hf._element.findall(f".//{{{NS}}}p")
                hf_changed += process_paragraphs(hf_paras, main_dict, kyuji_table)
            except Exception:
                pass

    # --- Font: Meiryo UI ---
    print(f"Setting font to {FONT_NAME} ...")
    set_font_all(doc.element.body, FONT_NAME)
    for section in doc.sections:
        for hf in [section.header, section.footer,
                   section.even_page_header, section.even_page_footer,
                   section.first_page_header, section.first_page_footer]:
            try:
                set_font_all(hf._element, FONT_NAME)
            except Exception:
                pass

    # --- Font size: 16pt digits -> 14pt ---
    print("Converting 16pt numeric runs to 14pt ...")
    sz_changed = convert_font_size(doc.element.body)
    for section in doc.sections:
        for hf in [section.header, section.footer,
                   section.even_page_header, section.even_page_footer,
                   section.first_page_header, section.first_page_footer]:
            try:
                sz_changed += convert_font_size(hf._element)
            except Exception:
                pass
    print(f"Font size changed: {sz_changed} run(s)")

    doc.save(output_path)
    print(f"Output: {output_path}")
    print(f"Changed: {body_changed} body para(s), {hf_changed} header/footer para(s)")

    # --- PDF conversion ---
    pdf_path = os.path.splitext(output_path)[0] + ".pdf"
    convert_to_pdf(output_path, pdf_path)
    print("Done.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    inp  = sys.argv[1]
    outp = sys.argv[2] if len(sys.argv) > 2 else None
    dic  = sys.argv[3] if len(sys.argv) > 3 else None
    convert(inp, outp, dic)
