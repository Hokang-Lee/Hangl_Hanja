# -*- coding: utf-8 -*-
"""
Korean to Hanja converter for Word (.docx) files
- Reads main dictionary from Excel Sheet1 (col A=Korean, col C=Hanja)
- Reads Kyuji->Shinji table from Excel Sheet2 (col A=old, col B=new)
- Handles cross-paragraph and soft-return splits
- Converts all fonts to Meiryo UI after conversion
- Condenses text inside parentheses () / （） to 95% width
- Subprocess-safe: opens/closes Word file via python-docx only

Usage:
    python convert_hanja.py "input.docx"
    python convert_hanja.py "output.docx"
    python convert_hanja.py "input.docx" "output.docx" "dict.xlsm"

Requirements:
    pip install python-docx lxml openpyxl
"""
import sys, os
import re as _re
from copy import deepcopy
from docx import Document
from lxml import etree

NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
FONT_NAME = "Meiryo UI"


# ---------------------------------------------------------------
# Console-safe printing
# ---------------------------------------------------------------
def safe_print(*args, sep=" ", end="\n", file=None, flush=False):
    if file is None:
        file = sys.stdout
    try:
        print(*args, sep=sep, end=end, file=file, flush=flush)
    except UnicodeEncodeError:
        try:
            enc = getattr(file, "encoding", None) or "utf-8"
            msg = sep.join(str(a) for a in args) + end
            file.buffer.write(msg.encode(enc, errors="replace"))
            if flush:
                file.flush()
        except Exception:
            pass


# ---------------------------------------------------------------
# Dictionary loader
# ---------------------------------------------------------------
def load_dicts(excel_path):
    try:
        import openpyxl
    except ImportError:
        safe_print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Sheet1: A=Korean, C=Hanja
    ws1 = wb.worksheets[0]
    main_dict = []
    for r in ws1.iter_rows(min_row=2, values_only=True):
        k = str(r[0]).strip() if r[0] else ""
        h = str(r[2]).strip() if r[2] else ""
        if k and h and k != h and k != "None" and h != "None":
            main_dict.append((k, h))
    main_dict.sort(key=lambda x: -len(x[0]))

    # Sheet2: A=旧字, B=当用漢字
    kyuji = []
    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
        for r in ws2.iter_rows(min_row=2, values_only=True):
            old = str(r[0]).strip() if r[0] else ""
            new = str(r[1]).strip() if r[1] else ""
            if old and new and old != new:
                kyuji.append((old, new))

    wb.close()
    safe_print(f"Main dict: {len(main_dict)} entries")
    safe_print(f"Kyuji table: {len(kyuji)} entries")
    return main_dict, kyuji


# ---------------------------------------------------------------
# XML helpers
# ---------------------------------------------------------------
def get_segments(para_elem):
    segs = []
    for child in para_elem.iter():
        tag = child.tag.split("}")[1] if "}" in child.tag else child.tag
        if tag == "t":
            segs.append([child.text or "", child, False])
        elif tag == "br":
            br_type = child.get(f"{{{NS}}}type", "textWrapping")
            if br_type in ("", "textWrapping"):
                segs.append(["\x0b", child, True])
    return segs


def apply_to_segments(segs, new_full):
    pos = 0
    for seg in segs:
        if seg[2]:
            pos += 1
            continue
        old_len = len(seg[0])
        seg[1].text = new_full[pos:pos + old_len]
        pos += old_len


# ---------------------------------------------------------------
# Replacement logic
# ---------------------------------------------------------------
def replace_text(text, dictionary):
    placeholders = {}
    result = text
    idx = 0
    for korean, hanja in dictionary:
        if korean in result:
            ph = f"\uE002{idx:04d}\uE003"
            placeholders[ph] = hanja
            result = result.replace(korean, ph)
            idx += 1
    for ph, hanja in placeholders.items():
        result = result.replace(ph, hanja)
    return result


def replace_kyuji(text, kyuji_table):
    for old, new in kyuji_table:
        text = text.replace(old, new)
    return text


def process_paragraphs(paras_elems, dictionary, kyuji_table):
    changed = 0

    para_segs = []
    para_texts = []
    for para in paras_elems:
        segs = get_segments(para)
        full = "".join(s[0] for s in segs)
        para_segs.append(segs)
        para_texts.append(full)

    # Step 1: Single-paragraph replacement
    new_texts = [replace_text(t, dictionary) for t in para_texts]

    # Step 2: Cross-paragraph replacement
    for i in range(len(para_texts) - 1):
        orig1 = para_texts[i]
        orig2 = para_texts[i + 1]
        combined_orig = orig1 + "\x00" + orig2
        for korean, hanja in dictionary:
            if len(korean) < 2:
                continue
            for j in range(1, len(korean)):
                part1 = korean[:j]
                part2 = korean[j:]
                split_k = part1 + "\x00" + part2
                if split_k not in combined_orig:
                    continue
                pos = combined_orig.find(split_k)
                boundary = len(orig1)
                if pos < boundary <= pos + len(split_k):
                    if pos > 0:
                        cb = combined_orig[pos - 1]
                        if ('가' <= cb <= '힣') or ('一' <= cb <= '鿿'):
                            continue
                    if orig1.endswith(part1) and orig2.startswith(part2):
                        hpart1 = hanja[:j] if j < len(hanja) else hanja
                        hpart2 = hanja[j:] if j < len(hanja) else ""
                        if new_texts[i].endswith(part1):
                            new_texts[i] = new_texts[i][:-len(part1)] + hpart1
                        converted_part2 = replace_text(part2, dictionary)
                        if new_texts[i + 1].startswith(converted_part2):
                            new_texts[i + 1] = hpart2 + new_texts[i + 1][len(converted_part2):]
                        elif new_texts[i + 1].startswith(part2):
                            new_texts[i + 1] = hpart2 + new_texts[i + 1][len(part2):]

    # Step 3: Kyuji -> Shinji
    new_texts = [replace_kyuji(t, kyuji_table) for t in new_texts]

    # Apply changes
    for segs, old, new in zip(para_segs, para_texts, new_texts):
        if old != new:
            apply_to_segments(segs, new)
            changed += 1

    return changed


# ---------------------------------------------------------------
# Run helpers
# ---------------------------------------------------------------
def _ensure_rpr(r):
    rpr = r.find(f"{{{NS}}}rPr")
    if rpr is None:
        rpr = etree.Element(f"{{{NS}}}rPr")
        r.insert(0, rpr)
    return rpr


def _get_run_text(r):
    t = r.find(f"{{{NS}}}t")
    return (t.text or "") if t is not None else ""


def _set_run_text(r, text):
    t = r.find(f"{{{NS}}}t")
    if t is None:
        t = etree.SubElement(r, f"{{{NS}}}t")
    if text.startswith(" ") or text.endswith(" "):
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    t.text = text


def _set_text_scale(rpr, percent):
    """
    Set character width scaling using w:w.
    100 = normal width, 95 = 95% width.
    """
    w_elem = rpr.find(f"{{{NS}}}w")
    if w_elem is None:
        w_elem = etree.Element(f"{{{NS}}}w")
        w_elem.set(f"{{{NS}}}val", str(percent))
        rpr.append(w_elem)
    else:
        w_elem.set(f"{{{NS}}}val", str(percent))


def _split_run_into_parts(para, run, parts):
    """
    Replace `run` in `para` with multiple runs according to `parts`.
    parts: list of dicts [{text: "...", scale: 95 or None}, ...]
      - scale None means keep original run properties
      - scale 95 means set w:w=95 for that part
    """
    children = list(para)
    try:
        idx = children.index(run)
    except ValueError:
        return 0

    inserted = 0
    for p in parts:
        txt = p["text"]
        if txt == "":
            continue
        new_run = deepcopy(run)
        _set_run_text(new_run, txt)
        if p.get("scale") is not None:
            rpr = _ensure_rpr(new_run)
            _set_text_scale(rpr, p["scale"])
        para.insert(idx + inserted, new_run)
        inserted += 1

    para.remove(run)
    return inserted


# ---------------------------------------------------------------
# Parentheses width condense
# ---------------------------------------------------------------
def condense_parentheses_width(elem, percent=95):
    """
    Condense text inside parentheses () / （） to given width percent.
    Parentheses themselves are included.
    Only the covered characters are condensed by splitting runs.
    """
    changed = 0

    for para in elem.iter(f"{{{NS}}}p"):
        runs = [r for r in para if r.tag == f"{{{NS}}}r"]
        if not runs:
            continue

        run_texts = [_get_run_text(r) for r in runs]
        full_text = "".join(run_texts)
        if not full_text:
            continue

        paren_ranges = [(m.start(), m.end()) for m in _re.finditer(r'[（(][^）)]*[）)]', full_text)]
        if not paren_ranges:
            continue

        inside = [False] * len(full_text)
        for a, b in paren_ranges:
            a = max(0, a)
            b = min(len(full_text), b)
            for k in range(a, b):
                inside[k] = True

        pos = 0
        run_positions = []
        for txt in run_texts:
            run_positions.append((pos, pos + len(txt)))
            pos += len(txt)

        for r, (rstart, rend) in list(zip(runs, run_positions)):
            if rstart == rend:
                continue
            if not any(inside[rstart:rend]):
                continue

            txt = _get_run_text(r)

            # all inside -> whole run condensed
            if all(inside[rstart:rend]):
                rpr = _ensure_rpr(r)
                _set_text_scale(rpr, percent)
                changed += 1
                continue

            # mixed -> split run
            parts = []
            seg_start = 0
            cur = inside[rstart]
            for i in range(1, len(txt)):
                flag = inside[rstart + i]
                if flag != cur:
                    parts.append({
                        "text": txt[seg_start:i],
                        "scale": percent if cur else None
                    })
                    seg_start = i
                    cur = flag
            parts.append({
                "text": txt[seg_start:],
                "scale": percent if cur else None
            })

            inserted = _split_run_into_parts(para, r, parts)
            if inserted > 0:
                for p in parts:
                    if p["text"] and p.get("scale") == percent:
                        changed += 1

    return changed


# ---------------------------------------------------------------
# Font conversion
# ---------------------------------------------------------------
def set_font_all(elem, font_name):
    """
    Safer font application:
    - touch run-level rPr only
    - do NOT modify paragraph-level pPr/rPr
    This helps preserve paragraph alignment / justification behavior.
    """
    for r in elem.iter(f"{{{NS}}}r"):
        rpr = r.find(f"{{{NS}}}rPr")
        if rpr is None:
            rpr = etree.Element(f"{{{NS}}}rPr")
            r.insert(0, rpr)

        for rf in rpr.findall(f"{{{NS}}}rFonts"):
            rpr.remove(rf)

        rf = etree.Element(f"{{{NS}}}rFonts")
        rf.set(f"{{{NS}}}ascii", font_name)
        rf.set(f"{{{NS}}}eastAsia", font_name)
        rf.set(f"{{{NS}}}hAnsi", font_name)
        rf.set(f"{{{NS}}}cs", font_name)
        rpr.insert(0, rf)


# ---------------------------------------------------------------
# PDF conversion
# ---------------------------------------------------------------
def convert_to_pdf(docx_path, pdf_path):
    try:
        from docx2pdf import convert as _pdf_convert
        safe_print("Converting to PDF (docx2pdf)...")
        _pdf_convert(docx_path, pdf_path)
        safe_print(f"PDF:    {pdf_path}")
        return
    except ImportError:
        pass
    except Exception as e:
        safe_print(f"docx2pdf failed: {e}")

    import subprocess, shutil
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice:
        try:
            safe_print("Converting to PDF (LibreOffice).")
            out_dir = os.path.dirname(os.path.abspath(pdf_path))
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf",
                 "--outdir", out_dir, docx_path],
                check=True, timeout=60
            )
            base = os.path.splitext(os.path.basename(docx_path))[0] + ".pdf"
            lo_out = os.path.join(out_dir, base)
            if lo_out != pdf_path and os.path.exists(lo_out):
                os.replace(lo_out, pdf_path)
            safe_print(f"PDF:    {pdf_path}")
            return
        except Exception as e:
            safe_print(f"LibreOffice failed: {e}")

    safe_print("WARNING: PDF conversion skipped.")
    safe_print("  To enable PDF output, run one of:")
    safe_print("    pip install docx2pdf       (requires Microsoft Word)")
    safe_print("    Install LibreOffice        (free)")


def find_dict_file(script_dir):
    candidates = []
    for name in os.listdir(script_dir):
        full = os.path.join(script_dir, name)
        if not os.path.isfile(full):
            continue
        low = name.lower()
        if low.endswith((".xlsx", ".xlsm")):
            candidates.append((0 if "convertalltext" in low else 1, full))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]


def convert(input_path, output_path=None, dict_path=None):
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = base + "_hanja" + ext

    if dict_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        dict_path = find_dict_file(script_dir)
        if dict_path is None:
            safe_print("ERROR: No Excel dictionary found in script folder.")
            sys.exit(1)

    safe_print(f"Dictionary: {dict_path}")
    main_dict, kyuji_table = load_dicts(dict_path)

    safe_print(f"Input:  {input_path}")
    doc = Document(input_path)

    # Body
    body_paras = doc.element.body.findall(f".//{{{NS}}}p")
    body_changed = process_paragraphs(body_paras, main_dict, kyuji_table)

    # Headers / Footers
    hf_changed = 0
    for section in doc.sections:
        for hf in [section.header, section.footer,
                   section.even_page_header, section.even_page_footer,
                   section.first_page_header, section.first_page_footer]:
            try:
                hf_paras = hf._element.findall(f".//{{{NS}}}p")
                hf_changed += process_paragraphs(hf_paras, main_dict, kyuji_table)
            except Exception:
                pass

    # Font unify
    safe_print(f"Setting font to {FONT_NAME} ...")
    set_font_all(doc.element.body, FONT_NAME)
    for section in doc.sections:
        for hf in [section.header, section.footer,
                   section.even_page_header, section.even_page_footer,
                   section.first_page_header, section.first_page_footer]:
            try:
                set_font_all(hf._element, FONT_NAME)
            except Exception:
                pass

    # Condense parentheses width
    safe_print("Condensing parentheses text width to 95% ...")
    width_changed = condense_parentheses_width(doc.element.body, percent=95)
    for section in doc.sections:
        for hf in [section.header, section.footer,
                   section.even_page_header, section.even_page_footer,
                   section.first_page_header, section.first_page_footer]:
            try:
                width_changed += condense_parentheses_width(hf._element, percent=95)
            except Exception:
                pass
    safe_print(f"Width changed: {width_changed} run/part(s)")

    doc.save(output_path)
    safe_print(f"Output: {output_path}")
    safe_print(f"Changed: {body_changed} body para(s), {hf_changed} header/footer para(s)")

    pdf_path = os.path.splitext(output_path)[0] + ".pdf"
    convert_to_pdf(output_path, pdf_path)
    safe_print("Done.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        safe_print(__doc__)
        sys.exit(1)
    inp = sys.argv[1]
    outp = sys.argv[2] if len(sys.argv) > 2 else None
    dic = sys.argv[3] if len(sys.argv) > 3 else None
    convert(inp, outp, dic)